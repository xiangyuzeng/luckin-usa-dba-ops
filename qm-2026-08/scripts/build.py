#!/usr/bin/env python3
"""QM 2026-08 data pack builder: raw/*.json -> ../*.xlsx + derived validation JSON."""
import json, re, collections, sys
from pathlib import Path
from datetime import datetime, timedelta
sys.path.insert(0, str(Path(__file__).resolve().parent))
from fs_kw import classify, is_low
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
RAW  = HERE.parent / "raw"
OUT  = HERE.parent
def L(n): return json.load(open(RAW / f"{n}.json", encoding="utf-8"))

# ---------- reference maps (labels only; every value list comes from the DB) --------
SEV = {1: "S 关键项 Critical", 2: "G 一般项 General", 3: "M 重点项 Major",
       4: "L 轻微项 Light", 9: "— 记录项 (0 分)"}
MODULE_CN = {
 'Cleaning and Sanitation': '清洁卫生', 'Process Control': '过程控制', 'Facility': '设施',
 'Document Record': '证照', 'Workplace Safety': '职业安全', 'Site Security': '职业安全',
 'Pests Control': '虫害防控', 'Maintenance of Equipment': '设备维护',
 'Approved Supplier': '供应链',
}
def module_cn(name):
    if not name: return ''
    n = name.strip()
    if n in MODULE_CN: return MODULE_CN[n]
    if n.startswith("Employees") and "Health" in n: return '员工健康卫生'
    if n.startswith("Temperature"): return '温控有效期'
    return f'(未映射) {n}'
PQNC_STATUS = {1:"1 已新建", 2:"2 已退回", 3:"3 待判责", 4:"4 待确认", 5:"5 已完成"}
PQNC_RESP   = {1:"1 (系统注释: 供应商)", 2:"2 (系统注释: 仓储)", 3:"3 (系统注释: 门店)",
               4:"4 供应商与仓储共担", 5:"5 责任不明", 6:"6 无责任 irresponsibility"}
PQNC_INIT   = {1:"门店", 2:"仓库"}
PQNC_PERIOD = {1:"收货时", 2:"存储时", 3:"使用时", 4:"售出后"}
PQNC_PROC   = {1:"无实物退货", 2:"实物退货"}
PQNC_JPARTY = {1:"采购", 2:"质控"}
RETURN_RSN  = {1:"属于产品正常现象", 2:"信息不充分", 3:"关闭不合格单"}
OPERATE     = {1:"判责", 2:"退回", 3:"确认", 4:"申诉"}
PROCESS_ST  = {10:"已完成", 20:"已生成", 30:"申诉中", 40:"整改中"}

def utc2ny(s):
    """UTC string -> America/New_York. Jul & Aug 2026 are both EDT (UTC-4); no DST edge."""
    if not s: return ''
    return (datetime.fromisoformat(str(s).replace('T', ' ')) - timedelta(hours=4)).strftime('%Y-%m-%d %H:%M:%S')

def parse_appeal(js):
    """t_shopcheck_opportunity.first/second_appeal_detail -> flat dict."""
    out = dict(submit_time='', submitter='', submitter_post='', describe='',
               result='', approve_time='', approver='', approver_post='', approve_remark='')
    if not js: return out
    try: d = json.loads(js)
    except Exception: return out
    op = d.get('operator') or {}
    out['describe']       = d.get('describe') or ''
    out['submit_time']    = (op.get('operateTime') or '')[:19].replace('T', ' ')
    out['submitter']      = op.get('empName') or ''
    out['submitter_post'] = op.get('empPost') or ''
    ar = d.get('approveResult')
    if isinstance(ar, dict):
        a = ar.get('approve')
        out['result'] = '通过(扣分归零)' if a == 1 else '不通过' if a == 0 else '待审批'
        out['approve_remark'] = ar.get('remark') or ''
        aop = ar.get('operator') or {}
        out['approve_time']  = (aop.get('operateTime') or '')[:19].replace('T', ' ')
        out['approver']      = aop.get('empName') or ''
        out['approver_post'] = aop.get('empPost') or ''
    else:
        out['result'] = '待审批'
    return out

def sheet(wb, title, rows, headers, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = re.sub(r'[\\\\/*?\\[\\]:]', '·', title)[:31]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(vertical='center')
    for r in rows: ws.append(r)
    ws.freeze_panes = 'A2'
    for i, h in enumerate(headers, 1):
        w = max(len(str(h)) * 1.6, *(min(len(str(r[i-1])), 48) for r in rows[:400])) if rows else len(str(h))*1.6
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 9), 52)
    return ws

# ================= shared frames ==================================================
stores    = L("stores")
LOCALITY  = {l['locality_mid']: l['locality_name'] for l in L("localities")}
store_by  = {s['dept_id']: s for s in stores}
V = {}   # validation results

def audit_frame(m):
    heads = {r['check_no']: r for r in L(f"audit_headers_{m}")}
    reps  = L(f"audit_reports_{m}")
    opps  = L(f"audit_opps_{m}")
    byrep = collections.defaultdict(list)
    for o in opps: byrep[o['check_no']].append(o)
    rows = []
    for r in reps:
        h  = heads.get(r['check_no'], {})
        os_ = [o for o in byrep.get(r['check_no'], []) if o['opp_deleted'] == 0]
        live = [o for o in os_ if o['opp_status'] == 1]
        starts_all  = [o['score_start'] for o in os_  if o['score_start'] is not None]
        starts_live = [o['score_start'] for o in live if o['score_start'] is not None]
        pre  = (min(starts_all)  if starts_all  else 100) + sum(o['original_score'] or 0 for o in os_)
        post = (min(starts_live) if starts_live else 100) + sum(o['original_score'] or 0 for o in live)
        rows.append(dict(report=r, head=h, opps=os_, pre=pre, post=post))
    return heads, rows, byrep, opps

PRIO = {'QA/QC-Store food safety audit': 0, 'OM-Area food safety Check': 1}
def main_inspection(frames):
    """主巡检: per store, QA > 区经 > 自检, then latest check_date, then largest check_no."""
    best = {}
    for f in frames:
        r = f['report']
        if f['head'].get('submitted') != 1 or f['head'].get('deleted') != 0: continue
        d = r['dept_id']
        k = (PRIO.get(r['large_category_name'], 2), r['check_date'], r['check_no'])
        if d not in best: best[d] = (k, f); continue
        ck = best[d][0]
        if k[0] < ck[0] or (k[0] == ck[0] and (k[1], k[2]) > (ck[1], ck[2])): best[d] = (k, f)
    return {d: v[1] for d, v in best.items()}

# ================= [1] 稽核报告主表 =================================================
H1 = ["是否主巡检","检查单号","检查大类(原始英文值)","检查大类ID","门店名称","门店编码(shop_no)","部门编码(dept_id)",
      "门店级别","所属城市","运营管理区域","检查日期","检查报告生成时间","检查人","检查人岗位编码",
      "检查途径(check_approach)","申诉前原始分","申诉后调整分","分差","机会点数(未删除)","已申诉归零数",
      "报告状态","已提交","已删除","流程状态","检查开始","检查结束","检查时长(秒)"]
def rows1(frames):
    global MAIN_IDS
    MAIN_IDS = {f['report']['check_no'] for f in main_inspection(frames).values()}
    out = []
    for f in sorted(frames, key=lambda x: (x['report']['check_date'], x['report']['check_no'])):
        r, h = f['report'], f['head']
        s = store_by.get(r['dept_id'], {})
        out.append([
            '是' if r['check_no'] in MAIN_IDS else '', r['check_no'], r['large_category_name'], r['large_category_id'],
            s.get('shop_name',''), s.get('shop_no',''), r['dept_id'],
            r.get('shop_level') or s.get('shop_level',''),
            LOCALITY.get(r.get('locality_mid'), r.get('locality_mid') or ''),
            h.get('operation_area') or '', str(r['check_date'])[:10],
            str(r.get('report_create_time') or '').replace('T',' '),
            r.get('checker_name',''), r.get('checker_post_code',''), r.get('check_approach'),
            f['pre'], f['post'], f['post']-f['pre'],
            len(f['opps']), sum(1 for o in f['opps'] if o['opp_status']==0),
            r.get('report_status'), h.get('submitted'), h.get('deleted'),
            PROCESS_ST.get(h.get('process_status'), h.get('process_status')),
            str(h.get('check_time_start') or '').replace('T',' '),
            str(h.get('check_time_end') or '').replace('T',' '), h.get('check_duration'),
        ])
    return out

heads8, fr8, byrep8, opps8 = audit_frame("2026-08")
heads7, fr7, byrep7, opps7 = audit_frame("2026-07")
wb = Workbook()
sheet(wb, "2026-08稽核报告", rows1(fr8), H1, first=True)
sheet(wb, "2026-07校验", rows1(fr7), H1)
# every LKUS check-header in the month, including非提交/已删除/测试大类 — nothing silently dropped
H1b = ["检查单号","检查大类(原始英文值)","检查大类ID","部门编码","门店名称","检查日期","已提交(status)",
       "已删除(deleted)","流程状态","检查人","是否生成报告"]
def rows1b(m, heads, byrep, reps_ids):
    out=[]
    for cid, h in sorted(heads.items(), key=lambda x: (x[1]['check_date'], x[0])):
        s = store_by.get(h['dept_id'], {})
        out.append([cid, h['large_category_name'], h['large_category_id'], h['dept_id'],
                    s.get('shop_name',''), str(h['check_date'])[:10], h['submitted'], h['deleted'],
                    PROCESS_ST.get(h.get('process_status'), h.get('process_status')),
                    h.get('checker_name',''), '是' if cid in reps_ids else '否'])
    return out
sheet(wb, "2026-08全部检查单(含未提交)", rows1b("2026-08", heads8, byrep8, {f['report']['check_no'] for f in fr8}), H1b)
sheet(wb, "2026-07全部检查单(含未提交)", rows1b("2026-07", heads7, byrep7, {f['report']['check_no'] for f in fr7}), H1b)
wb.save(OUT / "01_稽核报告主表_202608.xlsx")

# ================= [2] 机会点明细 ===================================================
H2 = ["检查单号","机会点ID","门店名称","部门编码","检查大类(原始英文值)","检查日期","模块(英文)","模块(中文)",
      "子项/标签","检查项全文","机会点描述","扣分类型码","扣分类型","原始机会点分值","当前机会点分值",
      "是否已归零","机会点状态(1有效/0无效)","机会点已删除","归零机制","一次申诉-提交时间","一次申诉-申请人",
      "一次申诉-申请岗位","一次申诉-理由","一次申诉-结果","一次申诉-审批时间","一次申诉-审批人",
      "一次申诉-审批岗位","一次申诉-审批意见","二次申诉-提交时间","二次申诉-结果","二次申诉-审批时间",
      "二次申诉-审批人","作废提交时间","作废审批时间","作废审批岗位","作废审批人","报告得分(申诉后)",
      "报告原始分(申诉前)","检查项ID","起评分(score_start)"]
discards = L("discard_all")
disc_by  = collections.defaultdict(list)
for d in discards: disc_by[d['shopcheck_data_id']].append(d)

def rows2(frames, opps):
    fmap = {f['report']['check_no']: f for f in frames}
    out = []
    for o in sorted(opps, key=lambda x: (x['check_date'], x['check_no'], x['opp_id'])):
        f  = fmap.get(o['check_no'])
        a1 = parse_appeal(o.get('first_appeal_detail'))
        a2 = parse_appeal(o.get('second_appeal_detail'))
        s  = store_by.get(o['dept_id'], {})
        zeroed = o['opp_status'] == 0
        orig = o['original_score']
        cur  = 0 if zeroed else orig
        mech = ''
        if zeroed:
            mech = '申诉审批通过' if (a1['result'].startswith('通过') or a2['result'].startswith('通过')) else '状态置为无效(无申诉记录)'
        ds = disc_by.get(o['check_no'], [])
        d0 = ds[-1] if ds else {}
        out.append([
            o['check_no'], o['opp_id'], s.get('shop_name',''), o['dept_id'], o['large_category_name'],
            str(o['check_date'])[:10], o.get('module_name') or '', module_cn(o.get('module_name')),
            o.get('leaf_name') or o.get('tag_name') or '', o.get('item_content') or '',
            o.get('opp_remark') or '', o.get('deduction_type'), SEV.get(o.get('deduction_type'),''),
            orig, cur, '是' if zeroed else '否', o['opp_status'], o['opp_deleted'], mech,
            a1['submit_time'], a1['submitter'], a1['submitter_post'], a1['describe'], a1['result'],
            a1['approve_time'], a1['approver'], a1['approver_post'], a1['approve_remark'],
            a2['submit_time'], a2['result'], a2['approve_time'], a2['approver'],
            str(d0.get('create_time') or '').replace('T',' '), str(d0.get('approve_time') or '').replace('T',' '),
            d0.get('applicant_post_code',''), d0.get('creator_name',''),
            f['post'] if f else '', f['pre'] if f else '', o['check_item_id'], o.get('score_start'),
        ])
    return out
wb = Workbook()
sheet(wb, "2026-08机会点明细", rows2(fr8, opps8), H2, first=True)
sheet(wb, "2026-07校验", rows2(fr7, opps7), H2)
sheet(wb, "作废流水(t_shopcheck_discard全表)",
      [[d['id'], d['tenant'], d['shopcheck_data_id'], d['approve'], d['reason'], d['applicant_post_code'],
        d['remark'], str(d['approve_time'] or '').replace('T',' '), d['creator_name'],
        str(d['create_time'] or '').replace('T',' ')] for d in discards],
      ["id","租户","检查单号","审批通过","作废原因","申请岗位","审批意见","审批时间","申请人","提交时间"])
wb.save(OUT / "02_稽核机会点明细_202608.xlsx")
print("[build] 01, 02 done")

# ================= [3] 检查项 <-> 模块映射 ==========================================
ic = L("item_config"); cc = L("category_config"); snap = L("config_snapshot_meta")
H3 = ["检查大类ID","检查大类(原始英文值)","模块ID","模块(英文)","模块(中文)","子项/叶子分类","检查项ID",
      "检查项全文","标签(tag)","扣分类型码","扣分类型(严重度)","标准分值(score_config)","起评分(score_start)",
      "排序","检查项状态","检查项已删除","检查项创建时间","检查项最后修改时间","叶子分类状态","分类树路径",
      "大类创建时间","大类最后修改时间"]
r3 = [[r['large_category_id'], r['large_category_name'], r['module_id'], r['module_name'],
       module_cn(r['module_name']), r['leaf_name'], r['check_item_id'], r['item_content'], r['tag_name'],
       r['deduction_type'], SEV.get(r['deduction_type'],''), r['score_config'], r['score_start'],
       r['sort'], r['item_status'], r['item_deleted'],
       str(r['item_create_time'] or '').replace('T',' '), str(r['item_modify_time'] or '').replace('T',' '),
       r['leaf_status'], r['leaf_path'], str(r['category_create_time'] or '').replace('T',' '),
       str(r['category_modify_time'] or '').replace('T',' ')]
      for r in ic if r['large_category_id']]
wb = Workbook(); sheet(wb, "检查项-模块映射", r3, H3, first=True)
sheet(wb, "分类树(全量)",
      [[c['id'], c['name'], c['parent_id'], c['path'], c['sort'], c['status'], c['deleted'],
        c['check_approach'], c['need_appeal'], c['appeal_approve_post_code'], c['need_improve'],
        c['allow_multiple_data'], c['discard_approve_post_code'],
        str(c['create_time'] or '').replace('T',' '), str(c['modify_time'] or '').replace('T',' ')] for c in cc],
      ["分类ID","名称","父级ID","路径","排序","状态","已删除","检查途径","需要申诉","申诉审批岗位","需要整改",
       "当月允许多单","作废审批岗位","创建时间","修改时间"])
sheet(wb, "模板快照(版本与生效期)",
      [[s['snapshot_id'], s['large_category_id'], str(s['snapshot_create_time'] or '').replace('T',' '),
        s['used_by_checks'], str(s['first_check'])[:10], str(s['last_check'])[:10]] for s in snap],
      ["快照ID(模板版本)","检查大类ID","快照生成时间","引用检查单数","最早检查日","最晚检查日"])
# 起评分规则 = 系统内的「-20」实现
prof = collections.Counter((r['deduction_type'], r['score_start']) for r in ic if r['item_deleted'] == 0)
sheet(wb, "评分规则验证",
      [[SEV.get(k[0],k[0]), k[0], k[1], v] for k, v in sorted(prof.items())],
      ["扣分类型","扣分类型码","起评分 score_start","检查项数量"])
wb.save(OUT / "03_检查项模块映射.xlsx")

# ================= [4] 门店主数据 ===================================================
H4 = ["是否在营门店","门店名称","门店编码(shop_no)","部门编码(dept_id)","部门名称","运营管理区域","门店级别","状态(status)",
      "开业日期","待停业日期","停业日期","是否封闭式","时区","国家","一级行政区","城市(二级行政区)",
      "三级行政区","地址","负责人","是否测试门店","创建时间","最后修改时间",
      "8月是否被稽核","8月稽核次数","7月稽核次数"]
aug_by_dept = collections.Counter(f['report']['dept_id'] for f in fr8
                                  if f['head'].get('submitted')==1 and f['head'].get('deleted')==0)
jul_by_dept = collections.Counter(f['report']['dept_id'] for f in fr7
                                  if f['head'].get('submitted')==1 and f['head'].get('deleted')==0)
r4 = [[('是' if (s['status']==1 and s['shop_level']=='SL02') else
        ('否-中央厨房/测试厨房' if s['shop_level']!='SL02' else '否-未开业')),
       s['shop_name'], s['shop_no'], s['dept_id'], s['dept_name'], s['operation_area'], s['shop_level'],
       s['status'], str(s['set_up_time'] or '')[:10], str(s['shut_up_time'] or '')[:10],
       str(s['off_time'] or '')[:10], s['close_type'], s['time_zone'], s['country_name'],
       s['administrative_area_name'], s['locality_name'], s['sublocality_name'], s['address'],
       s['manager_name'], s['test_flag'], str(s['create_time'] or '').replace('T',' '),
       str(s['modify_time'] or '').replace('T',' '),
       '是' if aug_by_dept.get(s['dept_id']) else '否',
       aug_by_dept.get(s['dept_id'], 0), jul_by_dept.get(s['dept_id'], 0)]
      for s in sorted(stores, key=lambda x: (x['status'], x['shop_no']))]
wb = Workbook(); sheet(wb, "门店主数据", r4, H4, first=True)
wb.save(OUT / "04_门店主数据.xlsx")
print("[build] 03, 04 done")

# ================= [5] PQNC 全量 ====================================================
sup_by  = {s['supplier_mid']: s for s in L("suppliers")}
spec_by = {g['spec_mid']: g for g in L("goods_spec")}
cell_by = {c['stock_cell_code']: c for c in L("stock_cells")}
unit_by = {u['unit_mid']: u['unit_name'] for u in L("units")}
loc_by  = {l['locality_mid']: l['locality_name'] for l in L("localities")}
ptype   = {}
for t in L("pqnc_type_cfg"):
    if t.get('language_code') == 'en-US' or t['pqnc_type_code'] not in ptype:
        ptype[t['pqnc_type_code']] = t.get('language_value') or t['name']

H5 = ["单号","单据状态码","单据状态","发起方","城市","库存单位名称","库存单位编号","运营管理区域",
      "当事人","发现问题时间(UTC)","发现问题时间(纽约)","发现问题时间段","货物大类","货物小类","规格名称",
      "规格编号(spec_mid)","批次号/生产日期","问题货物数量","数量单位","冷冻数量","冷藏数量","问题详细描述",
      "货值金额","核定金额","币种","待判责方","判责结果码","判责结果(系统注释)","判责说明","待确认方",
      "供应商名称","供应商ID","生产工厂","PQNC类型码(首次判责)","PQNC类型(首次判责)","PQNC类型码(最新判责)",
      "PQNC类型(最新判责)","是否食安(重建口径)","食安判定依据","发货单号","实物处理方式","问题货物留存",
      "异物留存","是否申诉过","是否自动确认","已退回次数","退回原因","新建时间(UTC)","新建时间(纽约)",
      "提交时间","判责时间","完成时间","最后修改时间","新建人","判责部门","删除标识",
      "首次判责码","首次判责结果","判责记录条数","是否被再次判责","首次判责时间","最后判责时间"]

SPOILAGE = re.compile(r'spoil|sour|rancid|rotten|curdl|smell|odor|discolor|mold|foreign|contaminat', re.I)
def pqnc_rows(tag, a_ny, b_ny):
    p_all = L(f"pqnc_{tag}")
    det   = collections.defaultdict(list)
    for d in L(f"pqnc_detail_{tag}"): det[d['pqnc_id']].append(d)
    rows, kept = [], []
    for p in p_all:
        ny = utc2ny(p['created_time'])
        if not (a_ny <= ny < b_ny): continue
        kept.append(p)
        ds  = sorted(det.get(p['pqnc_id'], []), key=lambda x: x['id'])
        jd  = [d for d in ds if d['operate_type'] == 1]
        rb  = [d for d in ds if d['operate_type'] == 2]
        t_first = jd[0]['one_pqnc_type_code'] if jd else None
        t_last  = jd[-1]['one_pqnc_type_code'] if jd else None
        jdesc   = '; '.join(d['description'] for d in jd if d.get('description'))
        cell = cell_by.get(p['stock_cell_code'], {})
        st   = store_by.get(cell.get('relate_dept_id')) or store_by.get(cell.get('storage_dept_id')) or {}
        sp   = spec_by.get(p['spec_mid'], {})
        # 食安 (rebuilt rule, see 99_取数说明与校验.md): Critical/Major judgment, or a
        # spoilage / foreign-object判责说明 or 问题描述.  Not a single system field.
        why = []
        if t_last in ('0002', '0003') or t_first in ('0002', '0003'): why.append('判责类型=Critical/Major')
        if SPOILAGE.search((p.get('problem_description') or '') + ' ' + jdesc): why.append('描述含变质/异物关键词')
        rows.append([
            p['pqnc_no'], p['status'], PQNC_STATUS.get(p['status'], p['status']),
            PQNC_INIT.get(p['initiator'], p['initiator']),
            loc_by.get(cell.get('locality_mid'), '') or st.get('locality_name', ''),
            cell.get('stock_cell_name', ''), p['stock_cell_code'], st.get('operation_area', ''),
            p['party_name'], str(p['discover_problems_time'] or '').replace('T',' '),
            utc2ny(p['discover_problems_time']),
            PQNC_PERIOD.get(p['discover_problems_time_period'], p['discover_problems_time_period']),
            sp.get('large_class_name',''), sp.get('small_class_name',''),
            sp.get('spec_name','') or sp.get('goods_name',''), p['spec_mid'],
            p['batch_no'], p['problem_goods_quantity'], unit_by.get(p['unit_mid'], p['unit_mid']),
            p['freeze_quantity'], p['refrigerate_quantity'], p['problem_description'],
            p['value_amount'], p['approved_amount'], p['settle_currency'],
            PQNC_JPARTY.get(p['party_of_judgment'], p['party_of_judgment']),
            p['responsibility'], PQNC_RESP.get(p['responsibility'], p['responsibility']), jdesc,
            PQNC_RESP.get(p['party_of_confirm'], p['party_of_confirm']),
            sup_by.get(p['supplier_mid'], {}).get('supplier_name', ''), p['supplier_mid'], p['factory_name'],
            t_first, ptype.get(t_first, ''), t_last, ptype.get(t_last, ''),
            '是' if why else '否', ' + '.join(why),
            p['ship_order_no'], PQNC_PROC.get(p['process_method'], p['process_method']),
            p['problem_goods_retention_status'], p['foreign_matter_retention_status'],
            p['appealed_flag'], p['auto_confirm_flag'], len(rb),
            '; '.join(RETURN_RSN.get(d['return_reason'], str(d['return_reason'])) for d in rb),
            str(p['created_time'] or '').replace('T',' '), ny,
            str(p['submit_time'] or '').replace('T',' '), str(p['judgment_time'] or '').replace('T',' '),
            str(p['complete_time'] or '').replace('T',' '), str(p['modified_time'] or '').replace('T',' '),
            p['creator_name'], p['judgment_dept_name'], p['delete_flag'],
            (jd[0]['responsibility'] if jd else None),
            PQNC_RESP.get(jd[0]['responsibility'], '') if jd else '',
            len(jd), '是' if len(jd) > 1 else '否',
            str(jd[0]['operated_time'] or '').replace('T', ' ') if jd else '',
            str(jd[-1]['operated_time'] or '').replace('T', ' ') if jd else '',
        ])
    return rows, kept, p_all

r5_aug, kept8, all8 = pqnc_rows("aug_wide", "2026-08-01 00:00:00", "2026-09-01 00:00:00")
r5_jul, kept7, all7 = pqnc_rows("jul_wide", "2026-07-01 00:00:00", "2026-08-01 00:00:00")
wb = Workbook()
sheet(wb, "2026-08 PQNC全量", r5_aug, H5, first=True)
sheet(wb, "2026-07校验", r5_jul, H5)
det8 = L("pqnc_detail_aug_wide"); id2no8 = {p['pqnc_id']: p['pqnc_no'] for p in all8}
sheet(wb, "2026-08操作流水",
      [[id2no8.get(d['pqnc_id'], d['pqnc_id']), d['operate_type'], OPERATE.get(d['operate_type'], ''),
        d['responsibility'], PQNC_RESP.get(d['responsibility'], ''), d['one_pqnc_type_code'],
        ptype.get(d['one_pqnc_type_code'], ''), d['two_pqnc_type_code'], d['three_pqnc_type_code'],
        RETURN_RSN.get(d['return_reason'], ''), d['description'], d['remarks'], d['operator_name'],
        d['operator_dept_name'], str(d['operated_time'] or '').replace('T',' ')]
       for d in det8 if d['pqnc_id'] in id2no8],
      ["单号","操作类型码","操作类型","判责码","判责结果","PQNC类型码","PQNC类型","二级类型","三级类型",
       "退回原因","说明","备注","操作人","操作部门","操作时间"])
# 判责重放: 已公布的 7 月数字对应「截至 2026-08-04 的判责状态」, 之后发生了大量再判责
det7 = L("pqnc_detail_jul_wide"); det7by = collections.defaultdict(list)
for d in det7: det7by[d['pqnc_id']].append(d)
def asof_resp(pid, asof):
    j = sorted([d for d in det7by[pid] if d['operate_type'] == 1
                and str(d['operated_time']).replace('T', ' ') <= asof], key=lambda x: x['id'])
    return j[-1]['responsibility'] if j else None
sheet(wb, "2026-07判责重放对照",
      [[p['pqnc_no'], p['value_amount'],
        asof_resp(p['pqnc_id'], '2026-08-04 00:00:00'),
        PQNC_RESP.get(asof_resp(p['pqnc_id'], '2026-08-04 00:00:00'), '未判责'),
        p['responsibility'], PQNC_RESP.get(p['responsibility'], '未判责'),
        '是' if asof_resp(p['pqnc_id'], '2026-08-04 00:00:00') != p['responsibility'] else '',
        str(p['modified_time'] or '').replace('T', ' ')]
       for p in kept7 if p['delete_flag'] == 0],
      ["单号","货值","判责码(截至2026-08-04)","判责结果(截至2026-08-04, = 已发布口径)","判责码(当前)",
       "判责结果(当前)","已变更","最后修改时间"])
sheet(wb, "PQNC类型字典", [[t['pqnc_type_code'], t['name'], t.get('language_code'), t.get('language_value'),
                            t['level'], t['weights'], t['status']] for t in L("pqnc_type_cfg")],
      ["类型码","名称","语言","多语言值","层级","权重","状态"])
wb.save(OUT / "05_PQNC全量_202608.xlsx")
print(f"[build] 05 done  aug={len(r5_aug)} jul={len(r5_jul)}")

# ================= [6] 客诉 · 食安类 =================================================
ord_by  = {o['order_id']: o for o in L("fs_orders")}
items   = collections.defaultdict(list)
for it in L("fs_order_items"): items[it['order_id']].append(it)
biz     = {b['code']: b for b in L("cs_biz_types")}
src     = {s['code']: s['name'] for s in L("cs_sources")}

H6 = ["投诉时间(纽约,到分钟)","投诉时间(UTC)","门店名称","门店编码","部门编码","渠道/来源","客诉原文",
      "分类(关键词判定)","置信度","关联订单ID","订单支付时间(纽约)","关联商品","评价等级(level)","评价标签",
      "是否投诉标记","客服是否回复","是否补偿发券","是否联系顾客","评价ID"]
def rows6(m):
    out = []
    for r in L(f"order_comments_{m}"):
        cats = classify(r['comment'])
        if not cats: continue
        o = ord_by.get(r['order_id'], {})
        s = store_by.get(o.get('shop_id'), {})
        low = is_low(r['comment'])
        out.append([
            utc2ny(r['create_time'])[:16], str(r['create_time']).replace('T',' '),
            o.get('shop_name','') or s.get('shop_name',''), s.get('shop_no',''), o.get('shop_id',''),
            'App 订单评价 (t_order_comment)', r['comment'], ' / '.join(cats),
            '低 (仅泛化措辞)' if low else '高', str(r['order_id']), utc2ny(o.get('pay_time')),
            ' | '.join(f"{i['spu_name']}({i.get('sku_name') or ''})" for i in items.get(r['order_id'], [])),
            r['level'], r['labels'], r['complaint_flag'], r['customer_reply_status'],
            r['compensation_send_coupon'], r['contact_customer'], str(r['id']),
        ])
    return sorted(out)
r6a, r6j = rows6("aug"), rows6("jul")
wb = Workbook()
sheet(wb, "2026-08食安类客诉候选", r6a, H6, first=True)
sheet(wb, "2026-07校验", r6j, H6)
# the ticket system, for completeness: LKUS food-safety branch is empty
def cs_rows(m):
    out = []
    for r in L(f"cs_sheets_{m}"):
        b = biz.get(r['l1_biz_type'], {})
        s = store_by.get(r['dept_id'], {})
        out.append([r['sheet_no'], utc2ny(r['create_time'])[:16], s.get('shop_name',''), r['dept_id'],
                    src.get(r['feedback_source'], r['feedback_source']), r['l1_biz_type'], b.get('name',''),
                    r['l2_biz_type'], biz.get(r['l2_biz_type'], {}).get('name',''), r['l3_biz_type'],
                    biz.get(r['l3_biz_type'], {}).get('name',''), r['status'],
                    (r['feedback_detail'] or '')[:400], (r['comment'] or '')[:400], r['has_compensated']])
    return out
H6b = ["工单号","创建时间(纽约)","门店","部门编码","反馈来源","一级类型码","一级类型","二级类型码","二级类型",
       "三级类型码","三级类型","工单状态","反馈详情","评价内容","是否赔付"]
sheet(wb, "2026-08客服工单(全量)", cs_rows("aug"), H6b)
sheet(wb, "2026-07客服工单(全量)", cs_rows("jul"), H6b)
sheet(wb, "客诉分类字典", [[b['code'], b['name'], b['parent_code'], b['level'], b['enabled']] for b in L("cs_biz_types")],
      ["编码","名称","父级编码","层级","启用"])
wb.save(OUT / "06_客诉食安类_202608.xlsx")

# ================= [7] 准入与供应商稽核 ==============================================
ent  = L("enterprise"); eq = L("enterprise_qual"); sq = L("supplier_qual")
sd   = L("spec_draft_2026"); sda = L("spec_draft_approved_2026")
sup  = L("suppliers"); glc = L("goods_large_class")
DRAFT_ST = {1:"1 已新建", 2:"2 已补全", 3:"3 已提交审批", 4:"4 审批通过", 5:"5 审批驳回", 6:"6 已关联主数据"}
wb = Workbook()
sheet(wb, "供应商准入-t_enterprise",
      [[e['enterprise_code'], e['enterprise_name'], e['tenant'], e['status'], e['sup_class'], e['nc_type'],
        e['business_scope'], e['submitter_name'], str(e['submit_time'] or '').replace('T',' '),
        e['first_auditor_name'], str(e['first_audit_time'] or '').replace('T',' '),
        e['second_auditor_name'], str(e['second_audit_time'] or '').replace('T',' '),
        str(e['created_time'] or '').replace('T',' '), e['reject_remark'], e['delete_flag']] for e in ent],
      ["企业编码","企业名称","租户","状态","供应商类型码","NC类型","经营范围","提交人","提交时间","一审人",
       "一审时间","二审人","二审时间","创建时间","驳回原因","删除标识"], first=True)
sheet(wb, "供应商主数据(2026新增)",
      [[s['supplier_mid'], s['supplier_name'], s['sup_class'], s['sup_nature'], s['nc_type'], s['status'],
        s['business_source_mid'], str(s['created_time'] or '').replace('T',' '), s['creator_name']]
       for s in sup if str(s['created_time'] or '') >= '2026-01-01'],
      ["供应商ID","供应商名称","供应商类型码","供应商形式","NC类型","状态","业务来源","创建时间","创建人"])
sheet(wb, "供应商资质",
      [[q['id'], q['supplier_id'], q['qualification_name'], str(q['validity_start'] or '')[:10],
        str(q['validity_end'] or '')[:10], q['validity_long_term_flag'], q['status'],
        str(q['modified_time'] or '').replace('T',' ')] for q in sq],
      ["id","供应商id","资质名称","有效期起","有效期止","长期有效","状态","修改时间"])
sheet(wb, "企业资质", [[q['id'], q['enterprise_id'], q['qualification_name'], str(q['validity_start'] or '')[:10],
                        str(q['validity_end'] or '')[:10], q['status'], q['tenant']] for q in eq],
      ["id","企业id","资质名称","有效期起","有效期止","状态","租户"])
sheet(wb, "原料准入-规格审批草稿",
      [[d['spec_mid'], d['spec_name'], d['goods_mid'], d['draft_status'], DRAFT_ST.get(d['draft_status'],''),
        d['current_approve_node'], str(d['created_time'] or '').replace('T',' '), d['creator_name'],
        str(d['modified_time'] or '').replace('T',' '), d['modifier_name']] for d in sd],
      ["规格编号","规格名称","货物编号","状态码","审批状态","当前审批节点","创建时间","创建人","修改时间","修改人"])
sheet(wb, "原料准入-2026通过或关联",
      [[d['spec_mid'], d['spec_name'], d['goods_mid'], DRAFT_ST.get(d['draft_status'],''),
        str(d['created_time'] or '').replace('T',' '), str(d['modified_time'] or '').replace('T',' '),
        d['modifier_name']] for d in sda],
      ["规格编号","规格名称","货物编号","审批状态","创建时间","通过/关联时间","操作人"])
sheet(wb, "货物大类(原料分类口径)", [[g['mid'], g['name'], g['status']] for g in glc], ["大类编号","名称","状态"])
sheet(wb, "供应商现场稽核", [["未找到承载表", "已对 64 台 MySQL 实例做 information_schema 全量扫描",
                              "匹配模式: audit / onsite / factory_check / verify / inspect",
                              "SRM 内无准入审核 / 年度审核 / 飞行检查记录表", "详见 99_取数说明与校验.md"]],
      ["结论","扫描范围","匹配模式","说明","出处"])
wb.save(OUT / "07_准入与供应商稽核_202608.xlsx")
print("[build] 06, 07 done")

# ================= [8] 月度趋势 2026-01 ~ 2026-08 ====================================
def cat_kind(name):
    """归类到报告使用的三类. 名称随模板版本变过 (1-6月无 QA/QC- 与 OM- 前缀), 因此按名称
    语义归类, 而不是把某一版的 large_category_id 写死."""
    n = (name or '')
    if 'self-check' in n:                    return '门店自检'
    if 'food safety audit' in n.lower():     return 'QA 稽核'
    if 'area food safety check' in n.lower():return '区经检查'
    return f'其他/测试: {n}'

# row-level headers for all 8 months -> exact per-month store coverage
hist = L("audit_headers_2026H1") + L("audit_headers_2026-07") + L("audit_headers_2026-08")
months = [f"2026-{i:02d}" for i in range(1, 9)]
agg = collections.defaultdict(lambda: [0, set()])
raw_rows = []
seen = collections.defaultdict(lambda: [0, set()])
for r in hist:
    ym = str(r['check_date'])[:7]
    kind = cat_kind(r['large_category_name'])
    valid = (r['submitted'] == 1 and r['deleted'] == 0)
    seen[(ym, r['large_category_id'], r['large_category_name'], r['submitted'], r['deleted'])][0] += 1
    seen[(ym, r['large_category_id'], r['large_category_name'], r['submitted'], r['deleted'])][1].add(r['dept_id'])
    if valid:
        agg[(ym, kind)][0] += 1
        agg[(ym, kind)][1].add(r['dept_id'])
for k, v in sorted(seen.items()):
    raw_rows.append([k[0], k[1], k[2], cat_kind(k[2]), k[3], k[4],
                     '计入' if (k[3] == 1 and k[4] == 0) else '不计入', v[0], len(v[1])])
kinds = ['门店自检', 'QA 稽核', '区经检查']
r8a = []
for ym in months:
    row = [ym]
    for k in kinds: row += [agg.get((ym, k), [0, set()])[0], len(agg.get((ym, k), [0, set()])[1])]
    row.append(sum(agg.get((ym, k), [0, set()])[0] for k in kinds))
    row.append(len(set().union(*[agg.get((ym, k), [0, set()])[1] for k in kinds])))
    r8a.append(row)
wb = Workbook()
sheet(wb, "稽核月度趋势", r8a,
      ["年月","门店自检 次数","门店自检 覆盖门店","QA稽核 次数","QA稽核 覆盖门店","区经检查 次数",
       "区经检查 覆盖门店","三类合计","当月被稽核门店数(去重)"], first=True)
sheet(wb, "稽核趋势-按原始大类", raw_rows,
      ["年月","检查大类ID","检查大类(原始英文值)","归类","已提交","已删除","是否计入","次数","覆盖门店"])

fj = L("trend_pqnc_firstjudge")
first_r = collections.defaultdict(collections.Counter); cur_r = collections.defaultdict(collections.Counter)
tot_n = collections.Counter(); tot_v = collections.defaultdict(float)
for r in fj:
    first_r[r['ym']][r['first_resp']] += r['n']; cur_r[r['ym']][r['current_resp']] += r['n']
    tot_n[r['ym']] += r['n']; tot_v[r['ym']] += float(r['value_amount'] or 0)
rej = {r['ym']: r['rejudged'] for r in L("trend_pqnc_rejudge")}
tp = L("trend_pqnc")
allst = collections.defaultdict(collections.Counter)
for r in tp: allst[r['ym']][(r['status'], r['delete_flag'])] += r['n']
r8b = [[ym, tot_n.get(ym, 0), round(tot_v.get(ym, 0), 2),
        first_r[ym].get(1, 0), cur_r[ym].get(1, 0),
        first_r[ym].get(2, 0), cur_r[ym].get(2, 0),
        first_r[ym].get(4, 0), cur_r[ym].get(4, 0),
        cur_r[ym].get(6, 0) + cur_r[ym].get(5, 0), rej.get(ym, 0),
        sum(n for (st, df), n in allst[ym].items() if df == 1)] for ym in months]
sheet(wb, "PQNC月度趋势", r8b,
      ["年月","PQNC 起数(未删除)","货值合计 USD","供应商责任(首次判责)","供应商责任(当前判责)",
       "仓储责任(首次判责)","仓储责任(当前判责)","共担(首次判责)","共担(当前判责)","未明确/无责任(当前)",
       "本月单中已被再次判责的单数","已删除单数"])
sheet(wb, "PQNC月度-状态分布",
      [[ym, PQNC_STATUS.get(st, st), df, n] for ym in months for (st, df), n in sorted(allst[ym].items())],
      ["年月","单据状态","删除标识","起数"])
tt = L("trend_pqnc_type")
sheet(wb, "PQNC月度-类型分布",
      [[r['ym'], r['one_pqnc_type_code'], ptype.get(r['one_pqnc_type_code'], '(无判责记录)'), r['n']] for r in tt],
      ["年月","类型码(首次判责)","类型","起数"])

# complaints trend — keyword-classified, both months pulled in full; other months = count only
tc = {r['ym']: r for r in L("trend_comments")}
cs = collections.defaultdict(collections.Counter)
for r in L("trend_cs"): cs[r['ym']][r['l1_biz_type']] += r['n']
sheet(wb, "客诉月度趋势",
      [[ym, tc.get(ym, {}).get('n', ''), sum(cs[ym].values()) or '', cs[ym].get('BIZ_TYPE_1', 0),
        len(r6j) if ym == '2026-07' else (len(r6a) if ym == '2026-08' else '未逐条判定')] for ym in months],
      ["年月","App订单评价数(有文字)","客服工单数","客服工单-Food safety 分类","食安类客诉(关键词候选)"])
ts = collections.defaultdict(collections.Counter)
for r in L("trend_suppliers"): ts[r['ym']][r['sup_class']] += r['n']
sheet(wb, "供应商主数据新增月度",
      [[ym, sum(ts[ym].values())] + [f"{k}:{v}" for k, v in sorted(ts[ym].items())] for ym in months],
      ["年月","新增供应商数(全类)","按供应商类型码分布 →"])
sheet(wb, "EHS工伤", [["取不到", "已对 64 台 MySQL 实例做 information_schema 全量扫描",
                       "匹配模式: ehs / injur / accident / incident / safety / casualty",
                       "无任何 EHS / 工伤事故承载表", "详见 99_取数说明与校验.md"]],
      ["结论","扫描范围","匹配模式","说明","出处"])
wb.save(OUT / "08_月度趋势_202601_202608.xlsx")
print("[build] 08 done")
json.dump({"aug_pqnc_rows": len(r5_aug), "jul_pqnc_rows": len(r5_jul),
           "aug_fs_complaints": len(r6a), "jul_fs_complaints": len(r6j)},
          open(RAW / "_build_summary.json", "w"), ensure_ascii=False, indent=1)
